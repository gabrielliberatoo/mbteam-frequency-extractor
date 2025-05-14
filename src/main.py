import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Mapeamento de meses para números
MONTHS = {
    'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
    'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
    'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'
}

def initialize_driver():
    """Inicializa e retorna o driver do Selenium configurado"""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-extensions")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def get_analysis_type():
    """Permite ao usuário escolher o tipo de análise"""
    print("\nEscolha o tipo de análise:")
    print("1. Por mês/ano (ex: maio 2025)")
    print("2. Por período personalizado (ex: 01/05/2025 a 31/05/2025)")
    
    while True:
        choice = input("Digite 1 ou 2: ")
        if choice in ('1', '2'):
            return choice
        print("Opção inválida. Tente novamente.")

def get_month_year():
    """Obtém o mês e ano para análise"""
    while True:
        try:
            month_year = input("\nDigite o mês e ano (ex: 'maio 2025'): ").lower().split()
            if len(month_year) != 2:
                raise ValueError
            
            month, year = month_year
            if month not in MONTHS:
                raise ValueError
                
            return month, year
            
        except ValueError:
            print("Formato inválido. Use 'mês ano' (ex: maio 2025)")

def get_custom_date_range():
    """Obtém um período personalizado do usuário"""
    print("\nPor favor, informe o período de análise (formato DD/MM/AAAA)")
    while True:
        try:
            start_date = input("Data inicial (DD/MM/AAAA): ")
            end_date = input("Data final (DD/MM/AAAA): ")
            
            # Validação das datas
            dt_start = datetime.strptime(start_date, "%d/%m/%Y")
            dt_end = datetime.strptime(end_date, "%d/%m/%Y")
            
            if dt_end < dt_start:
                print("A data final deve ser após a data inicial.")
                continue
                
            return start_date, end_date
            
        except ValueError:
            print("Formato de data inválido. Use DD/MM/AAAA.")

def get_student_links(driver):
    """Coleta todos os links para perfis de alunos na página atual"""
    try:
        student_cards = WebDriverWait(driver, 20).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li.list-group-item a.text-decoration-none[href*='/user/client/']"))
        )
        return list({card.get_attribute("href") for card in student_cards})
    except Exception as e:
        print(f"Erro ao coletar links de alunos: {e}")
        return []

def go_to_next_page(driver):
    """Navega para a próxima página de alunos"""
    try:
        next_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "ul.pagination li.page-item:not(.disabled) .fa-chevron-right"))
        )
        driver.execute_script("arguments[0].scrollIntoView();", next_button)
        time.sleep(1)
        next_button.click()
        time.sleep(2)
        return True
    except Exception as e:
        print(f"Não há mais páginas disponíveis: {e}")
        return False

def get_student_name(driver):
    """Extrai o nome do aluno"""
    try:
        return WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.fs-5"))
        ).text.strip()
    except Exception as e:
        print(f"Erro ao obter nome: {e}")
        return "Nome não encontrado"

def open_calendar_view(driver):
    """Abre a visualização de calendário do aluno"""
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.card.card-body.border-0.pt-2"))
        ).click()
        time.sleep(2)
        return True
    except Exception as e:
        print(f"Erro ao abrir calendário: {e}")
        return False

def navigate_to_month(driver, target_month, target_year):
    """Navega até o mês e ano especificados no calendário"""
    try:
        max_attempts = 24  # Máximo de 2 anos (12 meses para cada direção)
        attempts = 0
        
        while attempts < max_attempts:
            # Obtém o mês e ano atualmente exibidos - SELETORES ATUALIZADOS
            current_month_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.col-4.text-center span.text-capitalize")))
            current_year_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.col-4.text-center span.fw-bold.fs-5")))
            
            current_month = current_month_element.text.strip().lower()
            current_year = current_year_element.text.strip()
            
            # Verifica se chegamos no mês/ano desejado
            if current_month == target_month.lower() and current_year == target_year:
                return True
            
            # Converte para números para comparação
            current_month_num = MONTHS.get(current_month, 0)
            target_month_num = MONTHS.get(target_month.lower(), 0)
            
            if not current_month_num or not target_month_num:
                print("Mês não reconhecido no mapeamento")
                return False
                
            current_date_num = int(current_year) * 100 + int(current_month_num)
            target_date_num = int(target_year) * 100 + int(target_month_num)
            
            # Decide para qual direção navegar (só tem botão anterior)
            if target_date_num < current_date_num:
                # Navega para o mês anterior - SELETOR ATUALIZADO
                prev_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "button.no-style-btn.text-primary")))
                prev_button.click()
            else:
                # Não há botão "próximo" no HTML fornecido
                print("Não é possível navegar para meses futuros - o sistema só tem botão 'Anterior'")
                return False
            
            time.sleep(1.5)  # Aguarda o calendário atualizar
            attempts += 1
        
        print(f"Não foi possível navegar até {target_month}/{target_year} após {max_attempts} tentativas")
        return False
        
    except Exception as e:
        print(f"Erro ao navegar no calendário: {e}")
        return False

def get_monthly_summary(driver):
    """Obtém o resumo mensal de treinos, retorna 0 se não houver registros"""
    try:
        # Tentamos encontrar o elemento com um timeout reduzido
        summary_element = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.fw-semibold.fs-3.text-primary")))
        return int(summary_element.text)
    except:
        # Se não encontrar o elemento, assume 0 dias de treino
        return 0

def set_custom_date_range(driver, start_date, end_date):
    """Define um período personalizado no calendário"""
    try:
        # Clica para abrir o seletor de datas
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='Data inicial']"))
        ).click()
        time.sleep(1)
        
        # Preenche as datas
        start_field = driver.find_element(By.CSS_SELECTOR, "input[placeholder='Data inicial']")
        start_field.clear()
        start_field.send_keys(start_date)
        
        end_field = driver.find_element(By.CSS_SELECTOR, "input[placeholder='Data final']")
        end_field.clear()
        end_field.send_keys(end_date)
        
        # Aplica o filtro
        apply_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Aplicar')]"))
        )
        apply_button.click()
        time.sleep(3)
        return True
    except Exception as e:
        print(f"Erro ao definir período: {e}")
        return False

def count_custom_period_days(driver):
    """Conta dias de treino em período personalizado"""
    try:
        present_days = driver.find_elements(By.CSS_SELECTOR, "td.highlighted-day")
        return len(present_days)
    except Exception as e:
        print(f"Erro ao contar dias: {e}")
        return 0

def process_all_students(driver, start_date, end_date, analysis_type, target_month=None, target_year=None):
    """Processa todos os alunos conforme o tipo de análise"""
    all_students = []
    page_count = 1
    
    while True:
        print(f"\nProcessando página {page_count}...")
        student_links = get_student_links(driver)
        
        if not student_links:
            print("Nenhum aluno encontrado.")
            break
        
        for i, link in enumerate(student_links, 1):
            print(f"Aluno {i}/{len(student_links)}...", end=" ")
            
            driver.execute_script(f"window.open('{link}');")
            driver.switch_to.window(driver.window_handles[1])
            
            try:
                name = get_student_name(driver)
                print(f"{name}...", end=" ")
                
                if open_calendar_view(driver):
                    if analysis_type == '1':
                        if navigate_to_month(driver, target_month, target_year):
                            training_days = get_monthly_summary(driver)  # Já trata o caso 0 automaticamente
                            period = f"{target_month}/{target_year}"
                        else:
                            training_days = 0
                            period = "Erro de navegação"
                    else:
                        if set_custom_date_range(driver, start_date, end_date):
                            training_days = count_custom_period_days(driver)
                            period = f"{start_date} a {end_date}"
                        else:
                            training_days = 0
                            period = "Erro"
                    
                    all_students.append({
                        "Nome": name,
                        "Período": period,
                        "Dias de Treino": training_days
                    })
                    print(f"Dias: {training_days}")
                    
            except Exception as e:
                print(f"Erro: {str(e)[:50]}...")
            finally:
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                time.sleep(0.5)  # Tempo reduzido entre alunos
        
        if not go_to_next_page(driver):
            break
            
        page_count += 1
    
    return all_students

def save_to_excel(data, filename="Relatório de Frequência.xlsx"):
    """Salva os dados em Excel com formatação profissional"""
    if not data:
        print("Nenhum dado para salvar.")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Frequência"
    
    # Cabeçalhos
    headers = ["Aluno", "Período", "Dias de Treino"]
    ws.append(headers)
    
    # Estilo dos cabeçalhos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados
    for student in data:
        ws.append([student["Nome"], student["Período"], student["Dias de Treino"]])
    
    # Formatação
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal="center")
    
    # Ajuste de largura
    for col in ws.columns:
        max_length = max(
            len(str(cell.value)) for cell in col
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    
    wb.save(filename)
    print(f"\nRelatório salvo em: {os.path.abspath(filename)}")

def main():
    print("=== Coletor de Frequência MBTeam ===")
    driver = initialize_driver()
    
    try:
        # Configuração da análise
        analysis_type = get_analysis_type()
        
        if analysis_type == '1':
            target_month, target_year = get_month_year()
            start_date = f"01/{MONTHS[target_month.lower()]}/{target_year}"
            end_date = f"31/{MONTHS[target_month.lower()]}/{target_year}"
            period_label = f"mês {target_month}/{target_year}"
        else:
            start_date, end_date = get_custom_date_range()
            period_label = f"período {start_date} a {end_date}"
            target_month = target_year = None
        
        print(f"\nIniciando análise para {period_label}...")
        input("Pressione Enter após navegar para a lista de alunos...")
        
        start_time = time.time()
        student_data = process_all_students(
            driver, start_date, end_date, analysis_type,
            target_month, target_year
        )
        
        if student_data:
            save_to_excel(student_data)
            elapsed_time = time.time() - start_time
            print(f"\nConcluído em {elapsed_time:.2f}s | {len(student_data)} alunos processados")
        else:
            print("\nNenhum dado coletado.")
            
    except Exception as e:
        print(f"\nErro fatal: {e}")
    finally:
        driver.quit()
        print("\nFinalizado.")

if __name__ == "__main__":
    main()